import openpyxl

wb = openpyxl.load_workbook('_ExcelTest.xlsx')
sheet = wb['Sheet1']

print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=5, column=5).value)

print("%s" % sheet.max_row)
print("%s" % sheet.max_column)

# print("Hello, World!")
print("Cheers, Mate!")
